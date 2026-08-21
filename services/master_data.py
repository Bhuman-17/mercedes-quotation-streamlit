from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"
PRICE_LIST_FILE = REFERENCE_DIR / "Price List Compilation.xlsx"
RV_GRID_FILE = REFERENCE_DIR / "RV Grid.xlsx"


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except Exception:
        return None


def _rate(value: Any) -> float | None:
    number = _number(value)
    if number is None:
        return None
    if number > 1:
        number /= 100.0
    if number < 0 or number > 1:
        return None
    return round(number, 6)


def _mileage(value: Any) -> int | None:
    match = re.search(r"(\d[\d,]*)", _text(value))
    if not match:
        return None
    try:
        return int(match.group(1).replace(",", ""))
    except ValueError:
        return None


def _year_months(value: Any) -> int | None:
    match = re.search(r"(\d+)\s*yr", _text(value).lower())
    if not match:
        return None
    return int(match.group(1)) * 12


def _infer_powertrain(pt_code: Any, model_name: str, kufri_model: Any = None) -> str:
    pt = _text(pt_code).upper()
    combined = f"{_text(kufri_model)} {model_name}".strip()

    if pt in {"EQ", "EV", "BEV", "ELECTRIC"}:
        return "Electric"

    # Diesel variants in the pricing sheet are generally identified by a
    # trailing 'd' after the engine/variant number (220d, 300D, 450d, V300d...).
    if re.search(r"\b(?:[A-Za-z]*\s*)?\d{2,4}\s*[dD]\b", combined):
        return "Diesel"

    return "Petrol"


def _find_price_header(ws) -> tuple[int, dict[str, int], int]:
    """Return (header_row, key columns, latest ex-showroom column)."""
    for row in range(1, min(ws.max_row, 25) + 1):
        headers = {_norm(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1)}
        model_col = next((c for k, c in headers.items() if k == "model"), None)
        type_col = next((c for k, c in headers.items() if k == "typeclass"), None)
        pt_col = next((c for k, c in headers.items() if k == "pt"), None)
        if model_col and type_col and pt_col:
            ex_cols = [
                col
                for col in range(1, ws.max_column + 1)
                if _norm(ws.cell(row, col).value) in {"exshowroom", "exshowroomprice"}
            ]
            if not ex_cols:
                raise ValueError("No Ex-Showroom column found in Price List sheet.")
            return (
                row,
                {
                    "model": model_col,
                    "type_class": type_col,
                    "pt": pt_col,
                    "kufri": next((c for k, c in headers.items() if k == "kufrimodel"), 0),
                },
                max(ex_cols),
            )
    raise ValueError("Could not find the Price List header row (Model / Type Class / PT).")


def _load_price_models(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer the expected sheet, otherwise use the first non-empty sheet.
        ws = wb["Price List"] if "Price List" in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row, cols, price_col = _find_price_header(ws)

        models: dict[str, dict[str, Any]] = {}
        order: list[str] = []

        for row in range(header_row + 1, ws.max_row + 1):
            model_name = _text(ws.cell(row, cols["model"]).value)
            if not model_name:
                continue

            price = _number(ws.cell(row, price_col).value)
            if price is None or price <= 0:
                continue

            type_class = _text(ws.cell(row, cols["type_class"]).value)
            pt_code = _text(ws.cell(row, cols["pt"]).value)
            kufri = _text(ws.cell(row, cols["kufri"]).value) if cols["kufri"] else ""
            powertrain = _infer_powertrain(pt_code, model_name, kufri)

            key = model_name.casefold()
            if key not in models:
                models[key] = {
                    "name": model_name,
                    "powertrains": [],
                    "prices": {},
                    "typeClasses": {},
                    "ptCodes": {},
                    "kufriModels": {},
                }
                order.append(key)

            entry = models[key]
            if powertrain not in entry["powertrains"]:
                entry["powertrains"].append(powertrain)
            entry["prices"][powertrain] = round(price, 2)
            entry["typeClasses"][powertrain] = type_class
            entry["ptCodes"][powertrain] = pt_code
            entry["kufriModels"][powertrain] = kufri

        return [models[key] for key in order]
    finally:
        wb.close()


def _load_rv_grid(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["RV Grid"] if "RV Grid" in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = None
        car_col = None
        mileage_col = None
        for row in range(1, min(ws.max_row, 20) + 1):
            normalized = [_norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
            if "car" in normalized and any(v in {"avgmileage", "annualmileage"} for v in normalized):
                header_row = row
                car_col = normalized.index("car") + 1
                mileage_col = next(i + 1 for i, v in enumerate(normalized) if v in {"avgmileage", "annualmileage"})
                break

        if not header_row or not car_col or not mileage_col:
            raise ValueError("Could not find Car / Avg Mileage header in RV Grid.")

        # Read the year columns from the header itself. The current sheet uses
        # C:G for the primary/Petrol block and I:M for the Diesel block.
        year_columns: list[tuple[int, int]] = []
        for col in range(mileage_col + 1, ws.max_column + 1):
            months = _year_months(ws.cell(header_row, col).value)
            if months:
                year_columns.append((col, months))

        if not year_columns:
            raise ValueError("No 1yr-5yr columns found in RV Grid.")

        # Detect Petrol / Diesel group starting columns from the row above.
        group_row = max(1, header_row - 1)
        petrol_start = None
        diesel_start = None
        for col in range(1, ws.max_column + 1):
            label = _norm(ws.cell(group_row, col).value)
            if label == "petrol":
                petrol_start = col
            elif label == "diesel":
                diesel_start = col

        # Fallback for the known grid layout if group labels are moved/merged.
        if petrol_start is None:
            petrol_start = year_columns[0][0]

        primary_cols = [(c, m) for c, m in year_columns if diesel_start is None or c < diesel_start]
        diesel_cols = [(c, m) for c, m in year_columns if diesel_start is not None and c >= diesel_start]

        rv_models: dict[str, dict[str, Any]] = {}
        current_car = ""

        for row in range(header_row + 1, ws.max_row + 1):
            new_car = _text(ws.cell(row, car_col).value)
            mileage = _mileage(ws.cell(row, mileage_col).value)

            if new_car:
                current_car = re.sub(r"\s+", " ", new_car.replace("\n", " ")).strip()

            if not current_car or mileage is None:
                continue

            primary: dict[str, float] = {}
            diesel: dict[str, float] = {}

            for col, months in primary_cols:
                value = _rate(ws.cell(row, col).value)
                if value is not None:
                    primary[str(months)] = value

            for col, months in diesel_cols:
                value = _rate(ws.cell(row, col).value)
                if value is not None:
                    diesel[str(months)] = value

            # Ignore rows that carry no usable RV values.
            if not primary and not diesel:
                continue

            entry = rv_models.setdefault(
                current_car,
                {"primary": {}, "diesel": {}},
            )
            entry["primary"][str(mileage)] = primary
            if diesel:
                entry["diesel"][str(mileage)] = diesel

        return rv_models
    finally:
        wb.close()


def _platform_tokens(value: str) -> set[str]:
    text = _text(value).upper().replace("_", " ")
    return set(re.findall(r"[A-Z]\d{3}", text))


def _rv_platform_tokens(rv_name: str) -> set[str]:
    return _platform_tokens(rv_name)


def _compact_model(value: str) -> str:
    text = _norm(value)
    for token in (
        "mercedesbenz",
        "4matic",
        "amgline",
        "limitededition",
        "launchedition",
        "nightedition",
        "specialedition",
        "withnappaleather",
        "nappaleather",
        "seater",
        "performance",
        "new",
        "class",
    ):
        text = text.replace(token, "")
    return text


def _family_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", _text(value)).upper()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("COUPE", " ").replace("CABRIOLET", " ")
    text = re.sub(r"\b(4MATIC|AMG|LINE|LIMITED|LAUNCH|EDITION|HYBRID|PERFORMANCE|NEW|PRE|FL|S)\b", " ", text)
    tokens = re.findall(r"[A-Z]+|\d+", text)
    # Prefer a recognizable family token immediately before the first model number.
    for idx, token in enumerate(tokens):
        if token.isdigit() and idx > 0:
            family = tokens[idx - 1]
            if family not in {"E", "D"}:
                return family
    return tokens[0] if tokens else ""


def _choose_rv_model(
    model_name: str,
    type_class: str,
    powertrain: str,
    rv_models: dict[str, dict[str, Any]],
    model_alias: str = "",
) -> str | None:
    names = list(rv_models.keys())
    if not names:
        return None

    match_name = f"{model_name} {model_alias}".strip()
    lowered = match_name.lower()
    compact = _compact_model(match_name)

    # Electric products are named explicitly in the RV grid. These are
    # structural family aliases only; the rates themselves always come from
    # the workbook and are never hardcoded.
    if powertrain == "Electric":
        if "maybach" in lowered and "eqs" in lowered and "suv" in lowered:
            wanted = "eqssuvmm"
        elif "eqs" in lowered and "suv" in lowered:
            wanted = "eqssuv"
        elif "cla" in lowered:
            wanted = "clabev"
        elif "eqb" in lowered:
            wanted = "eqbfl"
        elif "eqa" in lowered:
            wanted = "eqa"
        elif "eqe" in lowered:
            wanted = "eqe"
        elif "eqs" in lowered:
            wanted = "eqseqs53"
        else:
            wanted = ""

        if wanted:
            for name in names:
                if _compact_model(name) == wanted or wanted in _compact_model(name):
                    return name

    # Current Maybach GLS pricing variants map to the dedicated GLS 600 MM
    # row when that row exists in the uploaded RV grid.
    if "maybach" in lowered and "gls" in lowered and "600" in lowered:
        for name in names:
            if _norm(name) == "gls600mm":
                return name

    # Prefer specific AMG/Dream Car rows before broad platform-family rows.
    # A candidate must match both the vehicle family (C/GLC/GLE/A/G/etc.)
    # and, when present, the model number. This prevents an AMG-line trim
    # from being mistaken for a dedicated AMG RV row.
    specific_names = [
        name
        for name in names
        if not _rv_platform_tokens(name) and _norm(name) != "allothermodels"
    ]

    model_family = _family_key(match_name)
    model_numbers = set(re.findall(r"\d{2,3}", match_name))
    best_name = None
    best_score = -1

    for name in specific_names:
        candidate_family = _family_key(name)
        candidate_numbers = set(re.findall(r"\d{2,3}", name))

        if candidate_family and model_family and candidate_family != model_family:
            continue
        if candidate_numbers and model_numbers and not (candidate_numbers & model_numbers):
            continue

        candidate = _compact_model(name)
        score = 0
        if candidate and (candidate in compact or compact in candidate):
            score += 20
        score += 6 if candidate_family and candidate_family == model_family else 0
        score += 8 * len(candidate_numbers & model_numbers)

        if score > best_score:
            best_name = name
            best_score = score

    if best_name and best_score >= 12:
        return best_name

    # CLE Cabriolet is named as a descriptive family row in the RV grid.
    if "cle" in lowered and "cabriolet" in lowered:
        for name in names:
            if "ccabrioletcle" == _norm(name):
                return name

    # Match normal cars by platform code. If the grid contains both original
    # and facelift rows for the same platform, prefer the FL row because the
    # pricing master is the current model master.
    type_tokens = _platform_tokens(type_class)
    type_digits = {re.sub(r"\D", "", token) for token in type_tokens if re.sub(r"\D", "", token)}
    platform_candidates: list[str] = []
    digit_candidates: list[str] = []
    for name in names:
        rv_tokens = _rv_platform_tokens(name)
        rv_digits = {re.sub(r"\D", "", token) for token in rv_tokens if re.sub(r"\D", "", token)}
        if type_tokens & rv_tokens:
            platform_candidates.append(name)
        elif type_digits & rv_digits:
            digit_candidates.append(name)

    # Only fall back to numeric platform matching when there is no exact
    # letter+number match (needed for compound labels such as W/V223).
    if not platform_candidates:
        platform_candidates = digit_candidates

    if platform_candidates:
        if len(platform_candidates) > 1:
            fl = [name for name in platform_candidates if re.search(r"\bFL\b", name, re.I)]
            if fl:
                return fl[0]
        return platform_candidates[0]

    # A few current products (for example V-Class) are most reliably matched
    # by a distinctive family word rather than a platform code in the grid.
    if model_name.upper().startswith("V300"):
        for name in names:
            if _norm(name) == "vclass":
                return name

    # Use the explicit catch-all row when the grid provides one.
    for name in names:
        if _norm(name) == "allothermodels":
            return name

    return None


def _rates_for_powertrain(rv_entry: dict[str, Any], powertrain: str) -> dict[str, dict[str, float]]:
    primary = rv_entry.get("primary") or {}
    diesel = rv_entry.get("diesel") or {}
    if powertrain == "Diesel" and diesel:
        return diesel
    # Electric/BEV rows in the supplied grid sit in the first (Petrol-side)
    # rate block; special Dream Car rows also use this single primary block.
    return primary


def load_master_data(
    price_path: Path = PRICE_LIST_FILE,
    rv_path: Path = RV_GRID_FILE,
) -> dict[str, Any]:
    if not price_path.exists():
        raise FileNotFoundError(f"Pricing master not found: {price_path}")
    if not rv_path.exists():
        raise FileNotFoundError(f"RV master not found: {rv_path}")

    price_models = _load_price_models(price_path)
    rv_models = _load_rv_grid(rv_path)

    for model in price_models:
        model["rvSource"] = {}
        model["rvRates"] = {}

        for powertrain in model["powertrains"]:
            source = _choose_rv_model(
                model["name"],
                model["typeClasses"].get(powertrain, ""),
                powertrain,
                rv_models,
                model["kufriModels"].get(powertrain, ""),
            )
            if source and source in rv_models:
                model["rvSource"][powertrain] = source
                model["rvRates"][powertrain] = _rates_for_powertrain(rv_models[source], powertrain)
            else:
                model["rvSource"][powertrain] = ""
                model["rvRates"][powertrain] = {}

    return {
        "models": price_models,
        "sourceFiles": {
            "priceList": price_path.name,
            "rvGrid": rv_path.name,
        },
    }


def enrich_inputs_from_master(inputs: dict[str, Any], master_data: dict[str, Any]) -> dict[str, Any]:
    """Server-side enforcement of price and Star Agility+ RV lookup."""
    enriched = dict(inputs)
    models = master_data.get("models") or []
    if not models:
        return enriched

    selected_name = _text(enriched.get("modelName"))
    model = next((m for m in models if m.get("name") == selected_name), None)
    if model is None:
        model = models[0]
        enriched["modelName"] = model.get("name", "")

    powertrains = model.get("powertrains") or []
    selected_powertrain = _text(enriched.get("powertrain"))
    if selected_powertrain not in powertrains:
        selected_powertrain = powertrains[0] if powertrains else ""
        enriched["powertrain"] = selected_powertrain

    prices = model.get("prices") or {}
    if selected_powertrain in prices:
        enriched["exShowroomPrice"] = prices[selected_powertrain]
    elif prices:
        enriched["exShowroomPrice"] = next(iter(prices.values()))

    mileage = str(int(_number(enriched.get("annualMileageAgility")) or 10000))
    tenure = str(int(_number(enriched.get("termAgility")) or 48))
    rates = ((model.get("rvRates") or {}).get(selected_powertrain) or {})
    rate = ((rates.get(mileage) or {}).get(tenure))
    enriched["rvPercentageAgility"] = float(rate) if rate is not None else 0.0

    return enriched
